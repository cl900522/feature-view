
# pip install openpyxl
import openpyxl
from openpyxl.styles import Font

print(openpyxl.__version__)


# wb = openpyxl.load_workbook("wb.xlsx")
wb = openpyxl.Workbook()
print(type(wb))

# sheet = wb.create_sheet("Sheet1")
sheet = wb.active

sheet.cell(row=1, column=1).value = 22
sheet.cell(row=1, column=2).value = "ChenMignxuan"

sheet.cell(row=2, column=1).value = 99
sheet.cell(row=2, column=2).value = "Renj"

sheet.cell(row=3, column=1).value = '=SUM(A1:A2)'

print("max_row:", sheet.max_row, "max_column:", sheet.max_column)

sheet.column_dimensions["A"].width = 20
sheet.column_dimensions["B"].width = 40
sheet.row_dimensions[1].height = 30
sheet["A1"].font = Font(sz=14, bold=True, italic=True)
sheet.cell(1, 2).font = Font(sz=14, bold=True)

print("max_row:", sheet.max_row, "max_column:", sheet.max_column)

wb.save("wb.xlsx")
