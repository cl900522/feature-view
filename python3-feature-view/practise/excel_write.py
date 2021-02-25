
# pip install openpyxl
import openpyxl
import datetime

from openpyxl.styles import Font, Alignment, numbers

print(openpyxl.__version__)

wb = openpyxl.Workbook()
print(type(wb))

# sheet = wb.create_sheet("Sheet1")
sheet = wb.active

rowIndex = 1
sheet.merge_cells(start_row=rowIndex, start_column=1,
                  end_row=rowIndex, end_column=2)
sheet.cell(rowIndex, 1).value = 'Title'
sheet.cell(rowIndex, 1).font = Font(
    sz=14, bold=True, italic=False)
sheet.cell(rowIndex, 1).alignment = Alignment(
    horizontal="center", vertical="center")
# sheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)


rows = 20
for i in range(rows):
    rowIndex = rowIndex+1
    sheet.cell(row=rowIndex, column=1).value = i
    sheet.cell(row=rowIndex, column=1).font = Font(
        sz=12, bold=False, italic=False)
    sheet.cell(row=rowIndex, column=2).value = ('JOIN{0}'.format(i))
    sheet.cell(row=rowIndex, column=2).font = Font(
        sz=12, bold=False, italic=True)
    sheet.cell(rowIndex, 3).value = datetime.datetime.now()
    sheet.cell(rowIndex, 4).value = datetime.datetime(2010, 7, 21)
    sheet.cell(rowIndex, 4).number_format = 'yyyy-mm-dd'

    sheet.cell(rowIndex, 5).value = 1231.123
    sheet.cell(rowIndex, 5).number_format = numbers.FORMAT_NUMBER
    sheet.cell(rowIndex, 6).value = 1231.123
    sheet.cell(rowIndex, 6).number_format = numbers.FORMAT_NUMBER_00

sheet.column_dimensions["A"].width = 10
sheet.column_dimensions["B"].width = 20
sheet.column_dimensions["C"].width = 20
sheet.column_dimensions["D"].width = 20

rowIndex = rowIndex+1
sheet.row_dimensions[rowIndex].height = 30
sheet.cell(row=rowIndex, column=1).value = '=SUM(A1:A{0})'.format(rows)
sheet.cell(rowIndex, 1).font = Font(sz=14, bold=True)

print("max_row:", sheet.max_row, "max_column:", sheet.max_column)

wb.save("wb.xlsx")
wb.close()
