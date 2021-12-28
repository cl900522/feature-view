
# pip install openpyxl
from re import A
import openpyxl
from openpyxl.styles import Font

print(openpyxl.__version__)

wb = openpyxl.load_workbook("C:/Users/jj/Desktop/未过账单据.xlsx")
sheet = wb.active
aMap = {}
bMap = {}

for i in range(10000):
    row = i+2
    orderId = sheet.cell(row=row, column=1).value
    storeId = sheet.cell(row=row, column=2).value

    if(orderId is None):
        break

    storeStr = "{}".format(storeId)
    storePart = storeId[-2:]
    list = aMap.get(storePart, [])
    list.append(orderId)
    aMap[storePart] = list

    storeSet = bMap.get(storePart, {})
    storeSet[storeId] = "1"
    bMap[storePart] = storeSet

wb.close()

for k in aMap.keys():
    print("update req_exec_record_{} set created=now(),call_time=now() where vendor_id=6342 and store_id in ({}) and ref_voucher_id in (\'{}\') and can_retry=1 and yn=1 and source='mq' and ref_voucher_type in ('ZFP','ZRE');".format(
        k, ','.join(bMap.get(k).keys()), "\',\'".join(aMap.get(k))))
